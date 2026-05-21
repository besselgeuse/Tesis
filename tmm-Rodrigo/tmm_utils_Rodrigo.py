
# -*- coding: utf-8 -*-
"""
Created on Thu Mar 26 17:45:17 2020

@author: simon

Utilities to calculate RAT of optical stacks with tmm_vec
"""
# 8/4 cambie los imports de Simon para que se pueda elegir la ruta del tmm_core

import numpy as np
import os
import pickle
import sys
import importlib.util
import torch
import torch.optim as optim

# 1. Definimos la ruta exacta al archivo tmm_core.py de Simon

tmm_path = r".\tmm_core.py"

# 2. Cargamos el módulo manualmente desde esa dirección
spec = importlib.util.spec_from_file_location("tmm", tmm_path)
tmm = importlib.util.module_from_spec(spec)
sys.modules["tmm"] = tmm # Esto asegura que otros scripts vean esta versión
spec.loader.exec_module(tmm)

from functools import partial
from itertools import product
from scipy.interpolate import interp1d
from openpyxl import load_workbook
import matplotlib.pyplot as plt


def cauchy_fn(A,B,C):
    def fn(lams):
        lams = lams*1.0
        return A+B*np.power(10,4)/np.power(lams,2)+C*np.power(10,9)/np.power(lams,4)
    return fn


def constant_fn(val):
    return partial(np.full_like, fill_value=val)

def n_eff(n1,n2,f):
    """
    Modelo de bruggeman. f es la fraccion de la especie 2.
    """
    if f == 0: return n1
    if f == 1: return n2
    e1 = n1**2
    e2 = n2**2
    #f = 1-f2
    omega = (1-f)*(e2-2*e1)+f*(e1-2*e2)
    return (np.sqrt(np.sqrt(omega**2 + 8*e1*e2)-omega))/2


def brugg_fn(n_a, n_b, f_b):
    def fn(lams):
        n1 = n_a(lams)
        n2 = n_b(lams)
        return n_eff(n1, n2, f_b)
    return fn

def n_eff_torch(n1, n2, f):
    e1 = n1**2
    e2 = n2**2
    # Ecuación de Bruggeman simétrica
    omega = (1-f)*(e2 - 2*e1) + f*(e1 - 2*e2)
    # sqrt de PyTorch sobre tensores complejos es nativa
    return torch.sqrt((torch.sqrt(omega**2 + 8*e1*e2) - omega) / 4.0)

def brugg_fn_torch(n_a, n_b, f_b):
    def fn(lams):
        # Asegúrate de que n_a y n_b devuelvan tensores PyTorch
        n1 = n_a(lams)
        n2 = n_b(lams)
        return n_eff_torch(n1, n2, f_b)
    return fn

def load_fn(filename, skiprows=1,delimiter = "\t"):
    
    wv, I = np.loadtxt(filename, skiprows=skiprows,delimiter=f"{delimiter}" ,unpack=True)
    return interp1d(wv, I,fill_value=(I[0],I[-1]), bounds_error=False)


def load_interp(filename, comments=';', skiprows=0, unit='nm'):
    '''loads optical data and outputs interpolation functions for n and k. 
    Output functions take wv in nm. Keep default arguments to load .nkv files
    '''
    scale = {'nm':1.0, 'um':1000.0}
    
    wv, n, k = np.loadtxt(filename, comments=comments, 
                      skiprows=skiprows, unpack=True)
    
    wv *= scale[unit]
    
    n_fn = interp1d(wv, n,fill_value=(n[0],n[-1]), bounds_error=False)
    k_fn = interp1d(wv, k,fill_value=(k[0],k[-1]), bounds_error=False)
    
    return n_fn, k_fn


def load_interp_in3(filename, unit='A'):
    '''loads optical data and outputs interpolation functions for n and k from .in3 files.
    Output functions take wv in nm. Default unit is Angstroms ('A').
    '''
    scale = {'nm': 1.0, 'um': 1000.0, 'A': 0.1}
    
    with open(filename, 'r') as f:
        lines = [line.strip() for line in f.readlines() if line.strip() != '']
    
    n_pts = 0
    k_pts = 0
    
    wv_n = []
    n_vals = []
    wv_k = []
    k_vals = []
    
    idx = 0
    # Buscar el número de puntos de n
    while idx < len(lines):
        parts = lines[idx].split()
        if len(parts) == 1:
            try:
                n_pts = int(parts[0])
                idx += 1
                break
            except ValueError:
                pass
        idx += 1
        
    for _ in range(n_pts):
        parts = lines[idx].split()
        wv_n.append(float(parts[0]))
        n_vals.append(float(parts[1]))
        idx += 1
        
    # Buscar el número de puntos de k
    while idx < len(lines):
        parts = lines[idx].split()
        if len(parts) == 1:
            try:
                k_pts = int(parts[0])
                idx += 1
                break
            except ValueError:
                pass
        idx += 1
        
    for _ in range(k_pts):
        parts = lines[idx].split()
        wv_k.append(float(parts[0]))
        k_vals.append(float(parts[1]))
        idx += 1
        
    wv_n = np.array(wv_n)
    wv_k = np.array(wv_k)
    n_vals = np.array(n_vals)
    k_vals = np.array(k_vals)
    
    if unit in scale:
        wv_n *= scale[unit]
        wv_k *= scale[unit]
        
    n_fn = interp1d(wv_n, n_vals, fill_value=(n_vals[0], n_vals[-1]), bounds_error=False)
    k_fn = interp1d(wv_k, k_vals, fill_value=(k_vals[0], k_vals[-1]), bounds_error=False)
    
    return n_fn, k_fn


def stack2tmm(stack, materials, lams, add_inf=True):
    '''Transforms a stack into three lists that can be fed into the 'tmm' package.
    BY DEFAULT ADDS inf AIR LAYERS'''
    n_list = []
    d_list = []
    c_list = []
    if add_inf:
        n_list.append([1.]*len(lams))
        d_list.append(np.inf)
        c_list.append('i')
        
    for layer in stack:
        d_list.append(layer[0])
        mat = materials[layer[1]]
        n_list.append(mat[0](lams) + mat[1](lams)*1.0j)
        c_list.append(layer[2])
        
    if add_inf:  
        n_list.append([1.]*len(lams))
        d_list.append(np.inf)
        c_list.append('i')
    
    return n_list, d_list, c_list

   
def RT_with_cache(file, stack, materials, lams, pol='s', th_0=0, thicks=None):
    if file is None:
        return calculate_RT(stack, materials, lams, pol, th_0, thicks)
    else:
    
        if os.path.exists(file):
            with open(file,'rb') as f:
                print('loading from cache\n',)
                RT = pickle.load(f)
        else:
            RT = calculate_RT(stack, materials, lams, pol, th_0, thicks)
            with open(file,'wb') as f:
                print('saving to cache\n')
                pickle.dump(RT, f)
        return RT


def calculate_RT(stack, materials, lams, pol='s', th_0=0, thicks=None):
    '''calulate R and T of a stack, for all combinations of thicks.'''
    n_list, d_list, c_list = stack2tmm(stack, materials, lams, add_inf=False)
    n_lams = len(lams)
    inc = 'i' in c_list[1:-1] #check if  any of the finite layers is incocherent
    if inc:
        def f(pol, n_list, d_list, th_0, lams, c_list):
            return tmm.inc_tmm(pol, n_list, d_list, c_list, th_0, lams)
    else:
        def f(pol, n_list, d_list, th_0, lams, c_list=None):
            return tmm.coh_tmm(pol, n_list, d_list, th_0, lams)
            
    if thicks is None:
        #case with only one thickness
        RAT = f(pol, n_list, d_list, th_0, lams, c_list)
        return RAT['R'], RAT['T']
        
    else:
        #case with varying thicknesses
        keys = thicks.keys()
        vals = thicks.values()
        sizes =  [len(val) for val in vals]
        params = product(*vals) #cartesian product of all thicknesses
        total = np.prod(sizes)
        # out = [np.nan]*total
        Rs = np.empty((total,n_lams))
        Ts = np.empty((total,n_lams))
        
        for i,par in enumerate(params):
            if not i%500: print(f'case {i}/{total}')
            
            for key,t in zip(keys, par):
                d_list[key] = t
            
            RAT = f(pol, n_list, d_list, th_0, lams, c_list)
            Rs[i] = RAT['R']
            Ts[i] = RAT['T']
            
        print('####finished####',)
        return Rs.reshape(*sizes,n_lams), Ts.reshape(*sizes,n_lams)

              



def calculate_RT_torch(n_list, d_bounds, lams, c_list=None, weights=None, pol='s', th_0=0.0, num_starts=50, num_epochs=150, lr=2.0, use_cuda=False):
    """
    Optimiza el espesor de un stack de capas delgadas usando PyTorch Autograd.
    
    Parámetros:
    - n_list: Tensor complejo de PyTorch (torch.complex128) con los índices [num_capas, num_lams].
    - d_bounds: Lista que define los límites o valores fijos para cada capa finita.
                Puede contener:
                - Tupla/Lista (min, max): para optimizar esa capa dentro de dichos límites.
                - Float/Int o Tupla (val, val): para mantener esa capa con un espesor constante fijo.
    - lams: Tensor con las longitudes de onda.
    - c_list: Lista opcional indicando 'c' (coherente) o 'i' (incoherente) para cada capa.
              Si se provee, utiliza la versión inc_tmm_torch para cálculos gruesos.
    - weights: Tensor opcional con los pesos (por ejemplo, espectro solar * IQE * wavelength)
               para realizar una optimización por reflectancia ponderada (Jsc).
    - pol: Polarización, 's' o 'p'.
    - th_0: Ángulo de incidencia (radianes).
    - num_starts: Cantidad de semillas aleatorias para evitar mínimos locales.
    - num_epochs: Número de iteraciones del Descenso de Gradiente.
    - lr: Learning rate (tasa de aprendizaje) para el optimizador Adam.
    - use_cuda: Si es True, utiliza la GPU para acelerar los cálculos (si está disponible).
    """
    device = torch.device("cuda" if use_cuda and torch.cuda.is_available() else "cpu")
    print(f"Ejecutando en dispositivo: {device}")
    
    # Mover tensores al dispositivo
    n_list = n_list.to(device)
    lams = lams.to(device)
    
    if weights is not None:
        if not isinstance(weights, torch.Tensor):
            weights = torch.tensor(weights, dtype=torch.float64, device=device)
        else:
            weights = weights.to(device)
        # Normalizar pesos para asegurar promedio ponderado correcto
        weights = weights / weights.sum()
    
    num_layers = n_list.shape[0]
    num_finite_layers = num_layers - 2
    num_wl = n_list.shape[1]
    
    if len(d_bounds) != num_finite_layers:
        raise ValueError(f"d_bounds debe tener longitud igual al número de capas finitas ({num_finite_layers}), pero tiene {len(d_bounds)}.")
    
    # 1. ANALIZAR CAPAS OPTIMIZABLES Y FIJAS
    is_optimizable = []
    opt_bounds = []  # Límites (min, max) para optimizables
    fixed_values = {}  # Mapa de idx -> valor_fijo para no optimizables
    
    for idx, b in enumerate(d_bounds):
        if isinstance(b, (tuple, list)):
            d_min, d_max = float(b[0]), float(b[1])
            if d_min == d_max:
                is_optimizable.append(False)
                fixed_values[idx] = d_min
            else:
                is_optimizable.append(True)
                opt_bounds.append((d_min, d_max))
        else:
            is_optimizable.append(False)
            fixed_values[idx] = float(b)
            
    num_opt_layers = sum(is_optimizable)
    print(f"Capas a optimizar: {num_opt_layers} | Capas fijas: {num_finite_layers - num_opt_layers}")
    
    if num_opt_layers == 0:
        raise ValueError("No se definieron capas optimizables en d_bounds (todas son fijas).")
        
    # 2. INICIALIZACIÓN MULTI-START EN EL ESPACIO NO ACOTODO (LOGIT)
    # Inicializamos d_initial en el espacio físico dentro de [d_min, d_max],
    # y luego lo mapeamos a d_opt no acotado para usar el truco del Sigmoide.
    d_initial = torch.zeros((num_starts, num_opt_layers), dtype=torch.float64, device=device)
    opt_idx = 0
    for idx, opt in enumerate(is_optimizable):
        if opt:
            d_min, d_max = opt_bounds[opt_idx]
            d_init_phys = torch.empty(num_starts, device=device).uniform_(d_min, d_max)
            # Mapeo inverso de sigmoide (logit)
            p = (d_init_phys - d_min) / (d_max - d_min)
            p = torch.clamp(p, 1e-7, 1.0 - 1e-7)  # Evitar desbordes numéricos
            d_initial[:, opt_idx] = torch.log(p / (1.0 - p))
            opt_idx += 1
            
    d_opt = d_initial.clone().detach().requires_grad_(True)
    
    # 3. OPTIMIZADOR
    optimizer = optim.Adam([d_opt], lr=lr)
    
    print(f"Iniciando optimización con Autograd ({num_starts} semillas en paralelo)...")
    
    # Columna de infinitos para sustrato/superestrato: [num_starts, 1]
    inf_col = torch.full((num_starts, 1), float('inf'), dtype=torch.float64, device=device)
    
    # Helper para construir d_fisico a partir de d_opt y las capas fijas
    def reconstruct_d_fisico(d_opt_tensor):
        d_fisico_cols = []
        o_idx = 0
        for idx, opt in enumerate(is_optimizable):
            if opt:
                d_min, d_max = opt_bounds[o_idx]
                # Mapeo sigmoide: garantiza d_min <= espesor <= d_max estrictamente
                col = d_min + (d_max - d_min) * torch.sigmoid(d_opt_tensor[:, o_idx])
                d_fisico_cols.append(col)
                o_idx += 1
            else:
                val = fixed_values[idx]
                col = torch.full((d_opt_tensor.shape[0],), val, dtype=torch.float64, device=device)
                d_fisico_cols.append(col)
        return torch.stack(d_fisico_cols, dim=1)
    
    # 4. BUCLE DE OPTIMIZACIÓN (sin loop sobre semillas)
    for epoch in range(num_epochs):
        optimizer.zero_grad()
        
        d_fisico = reconstruct_d_fisico(d_opt)  # [num_starts, num_finite_layers]
        
        # Construir d_list BATCHED: [num_starts, num_layers]
        d_full = torch.cat([inf_col, d_fisico, inf_col], dim=1)
        # Expandir a [num_starts, num_layers, num_wl]
        d_full_3d = d_full.unsqueeze(-1).expand(-1, -1, num_wl)
        
        # UNA SOLA llamada para TODAS las semillas
        if c_list is None:
            result = tmm.coh_tmm_torch(pol=pol, n_list=n_list, d_list=d_full_3d,
                                        th_0=th_0, lam_vac=lams)
        else:
            result = tmm.inc_tmm_torch(pol=pol, n_list=n_list, d_list=d_full_3d, c_list=c_list,
                                        th_0=th_0, lam_vac=lams)
            
        R = result['R']  # [num_starts, num_wl]
        
        # Loss (promedio simple o ponderado por weights)
        if weights is None:
            loss_per_start = R.mean(dim=-1)
        else:
            loss_per_start = (R * weights).sum(dim=-1)
            
        loss = loss_per_start.sum()
        
        loss.backward()
        optimizer.step()
        
        if epoch % 50 == 0:
            mejor_R_actual = loss_per_start.min().item()
            if weights is None:
                print(f"Epoch {epoch}/{num_epochs} | Mejor Reflectancia Media: {mejor_R_actual*100:.2f}%")
            else:
                print(f"Epoch {epoch}/{num_epochs} | Mejor Reflectancia Media Ponderada: {mejor_R_actual*100:.2f}%")

    # 5. EXTRACCIÓN DEL MÍNIMO GLOBAL
    with torch.no_grad():
        d_final_fisico = reconstruct_d_fisico(d_opt)
        d_full = torch.cat([inf_col, d_final_fisico, inf_col], dim=1)
        d_full_3d = d_full.unsqueeze(-1).expand(-1, -1, num_wl)
        if c_list is None:
            result = tmm.coh_tmm_torch(pol=pol, n_list=n_list, d_list=d_full_3d,
                                        th_0=th_0, lam_vac=lams)
        else:
            result = tmm.inc_tmm_torch(pol=pol, n_list=n_list, d_list=d_full_3d, c_list=c_list,
                                        th_0=th_0, lam_vac=lams)
        R_final = result['R']
    
    if weights is None:
        loss_final = R_final.mean(dim=-1)
    else:
        loss_final = (R_final * weights).sum(dim=-1)
        
    best_idx = loss_final.argmin()
    
    best_thicknesses = d_final_fisico[best_idx].cpu().detach().numpy()
    best_R_curve = R_final[best_idx].cpu().detach().numpy()
    
    print(f"\n¡Optimización completada!")
    print(f"Espesores óptimos encontrados: {best_thicknesses} nm")
    if weights is None:
        print(f"Reflectancia media óptima: {loss_final[best_idx].item()*100:.2f}%")
    else:
        print(f"Reflectancia media ponderada óptima: {loss_final[best_idx].item()*100:.2f}%")
    
    return best_thicknesses, best_R_curve


def coh_tmm_torch_batched(pol, n_list, d_list, th_0, lam_vac):
    """
    Versión diferenciable y paralelizada (doblemente BATCHED: por semillas y lambdas) 
    del Método de la Matriz de Transferencia en PyTorch.
    Soporta que n_list y d_list tengan forma [batch, num_layers, num_wl].
    """
    device = n_list.device
    batch_size, num_layers, num_wl = n_list.shape
    
    if not isinstance(lam_vac, torch.Tensor):
        lam_vac = torch.tensor(lam_vac, dtype=torch.float64, device=device)
    else:
        lam_vac = lam_vac.to(device)
        
    if not isinstance(th_0, torch.Tensor):
        th_0 = torch.tensor(th_0, dtype=n_list.real.dtype, device=device)
    else:
        th_0 = th_0.to(device)
        
    sin_th_0 = torch.sin(th_0)
    # th_list shape: [batch, num_layers, num_wl]
    # Slicing n_list[:, 0:1, :] obtiene el superestrato (capa 0) para cada lote
    th_list = torch.asin(n_list[:, 0:1, :] * sin_th_0 / n_list)
    
    # kz_list: [batch, num_layers, num_wl]
    kz_list = 2 * torch.pi * n_list * torch.cos(th_list) / lam_vac.view(1, 1, num_wl)
    
    # delta: [batch, num_layers, num_wl]
    delta = kz_list * d_list
    
    # Prevenir desbordamiento por absorción alta en capas internas finitas
    inner = delta[:, 1:-1, :]
    cond = inner.imag > 100
    inner = torch.where(cond, inner.real + 100j, inner)
    delta = torch.cat([delta[:, :1, :], inner, delta[:, -1:, :]], dim=1)
    
    t_list_vals = []
    r_list_vals = []
    
    for i in range(num_layers - 1):
        n_i = n_list[:, i, :]
        n_f = n_list[:, i+1, :]
        th_i = th_list[:, i, :]
        th_f = th_list[:, i+1, :]
        
        cos_th_i = torch.cos(th_i)
        cos_th_f = torch.cos(th_f)
        
        if pol == 's':
            r_val = (n_i * cos_th_i - n_f * cos_th_f) / (n_i * cos_th_i + n_f * cos_th_f)
            t_val = 2 * n_i * cos_th_i / (n_i * cos_th_i + n_f * cos_th_f)
        elif pol == 'p':
            r_val = (n_f * cos_th_i - n_i * cos_th_f) / (n_f * cos_th_i + n_i * cos_th_f)
            t_val = 2 * n_i * cos_th_i / (n_f * cos_th_i + n_i * cos_th_f)
        else:
            raise ValueError("Polarization must be 's' or 'p'")
            
        r_list_vals.append(r_val)
        t_list_vals.append(t_val)
        
    # Inicializar la matriz de transferencia Mtilde = [batch, num_wl, 2, 2]
    ones_bw = torch.ones(batch_size, num_wl, dtype=delta.dtype, device=device)
    zeros_bw = torch.zeros_like(ones_bw)
    
    Mtilde = torch.stack([
        torch.stack([ones_bw, zeros_bw], dim=-1),
        torch.stack([zeros_bw, ones_bw], dim=-1)
    ], dim=-2)
    
    for i in range(1, num_layers - 1):
        delta_i = delta[:, i, :]
        r_i = r_list_vals[i]
        t_i = t_list_vals[i]
        
        exp_minus = torch.exp(-1j * delta_i)
        exp_plus = torch.exp(1j * delta_i)
        zeros_i = torch.zeros_like(delta_i)
        ones_i = torch.ones_like(delta_i)
        
        # A y B shape: [batch, num_wl, 2, 2]
        A = torch.stack([
            torch.stack([exp_minus, zeros_i], dim=-1),
            torch.stack([zeros_i, exp_plus], dim=-1)
        ], dim=-2)
        
        B = torch.stack([
            torch.stack([ones_i, r_i], dim=-1),
            torch.stack([r_i, ones_i], dim=-1)
        ], dim=-2)
        
        d_coeff = (1.0 / t_i).unsqueeze(-1).unsqueeze(-1)
        M_i = d_coeff * torch.matmul(A, B)
        Mtilde = torch.matmul(Mtilde, M_i)
        
    # Interfaz frontal (0 -> 1)
    r_0 = r_list_vals[0]
    t_0 = t_list_vals[0]
    ones_0 = torch.ones_like(r_0)
    
    A_front = torch.stack([
        torch.stack([ones_0, r_0], dim=-1),
        torch.stack([r_0, ones_0], dim=-1)
    ], dim=-2)
    
    d_front = (1.0 / t_0).unsqueeze(-1).unsqueeze(-1)
    M_front = d_front * A_front
    Mtilde = torch.matmul(M_front, Mtilde)
    
    r = Mtilde[:, :, 1, 0] / Mtilde[:, :, 0, 0]
    t = 1.0 / Mtilde[:, :, 0, 0]
    
    return {'r': r, 't': t}


def fit_ellipsometry_torch(n_list, d_bounds, lams, Is_exp, Ic_exp, c_list=None, th_0=0.0, 
                           num_starts=50, num_epochs=150, lr=2.0, use_cuda=False, 
                           layer_models=None, layer_names=None):
    """
    Optimiza el espesor de un stack de capas delgadas y opcionalmente los parámetros 
    de modelos de dispersión (como Cauchy) usando PyTorch Autograd.
    
    Parámetros:
    - n_list: Tensor complejo de PyTorch con los índices [num_capas, num_lams] o lista de índices/tensores.
    - d_bounds: Lista de límites para los espesores de capas finitas.
    - lams: Tensor con las longitudes de onda.
    - Is_exp: Tensor PyTorch con Is experimental.
    - Ic_exp: Tensor PyTorch con Ic experimental.
    - c_list: Ignorado (mantenido por compatibilidad).
    - th_0: Ángulo de incidencia en radianes.
    - num_starts: Cantidad de semillas aleatorias paralelas.
    - num_epochs: Número de épocas del optimizador.
    - lr: Tasa de aprendizaje.
    - use_cuda: Usar GPU si está disponible.
    - layer_models: Lista de modelos de dispersión por capa. e.g. [None, 'cauchy', None].
      Soporta 'cauchy', 'cauchy_absorbent' o un diccionario con 'model', 'bounds', 'initial'.
    - layer_names: Lista opcional con los nombres de las capas.
    """
    device = torch.device("cuda" if use_cuda and torch.cuda.is_available() else "cpu")
    print(f"Ejecutando en dispositivo: {device}")
    
    if c_list is not None:
        print("ADVERTENCIA: c_list fue proporcionado pero será ignorado en elipsometría.")
    
    # Asegurar que lams, Is_exp, Ic_exp estén en el dispositivo
    lams = lams.to(device)
    
    if not isinstance(Is_exp, torch.Tensor):
        Is_exp = torch.tensor(Is_exp, dtype=torch.float64, device=device)
    else:
        Is_exp = Is_exp.to(device)
        
    if not isinstance(Ic_exp, torch.Tensor):
        Ic_exp = torch.tensor(Ic_exp, dtype=torch.float64, device=device)
    else:
        Ic_exp = Ic_exp.to(device)
        
    # Convertir n_list a tensor en dispositivo si es un tensor, o lista de tensores
    if isinstance(n_list, torch.Tensor):
        n_list = n_list.to(device)
        num_layers = n_list.shape[0]
        num_wl = n_list.shape[1]
    else:
        # Si es una lista, convertir tensores individuales
        n_list_tensors = []
        for item in n_list:
            if isinstance(item, torch.Tensor):
                n_list_tensors.append(item.to(device))
            else:
                # Es un dummy o placeholder para material paramétrico
                n_list_tensors.append(torch.ones_like(lams, dtype=torch.complex128, device=device))
        n_list = torch.stack(n_list_tensors, dim=0)
        num_layers = n_list.shape[0]
        num_wl = n_list.shape[1]
        
    num_finite_layers = num_layers - 2
    
    if len(d_bounds) != num_finite_layers:
        raise ValueError(f"d_bounds debe tener longitud {num_finite_layers}, pero tiene {len(d_bounds)}.")
        
    # 1. ANALIZAR MODELOS DE DISPERSIÓN PARAMÉTRICOS
    is_parametric = False
    disp_layers = []
    total_disp_params = 0
    
    if layer_models is not None:
        for idx, model in enumerate(layer_models):
            if model is not None:
                is_parametric = True
                
                if isinstance(model, dict):
                    model_type = model.get('model', 'cauchy')
                    bounds = model.get('bounds', None)
                    initial = model.get('initial', None)
                else:
                    model_type = model
                    bounds = None
                    initial = None
                
                # Definir cantidad de parámetros y valores por defecto
                if model_type == 'cauchy':
                    num_p = 3
                    # A en [1.0, 3.0], B y C en [-1.0, 1.0]
                    default_bounds = [(1.0, 3.0), (-1.0, 1.0), (-1.0, 1.0)]
                    default_initial = [1.5, 0.0, 0.0]
                elif model_type == 'cauchy_absorbent':
                    num_p = 6
                    # A en [1.0, 3.0], B y C en [-1.0, 1.0], D en [0.0, 2.0], E y F en [-1.0, 1.0]
                    default_bounds = [(1.0, 3.0), (-1.0, 1.0), (-1.0, 1.0), (0.0, 2.0), (-1.0, 1.0), (-1.0, 1.0)]
                    default_initial = [1.5, 0.0, 0.0, 0.0, 0.0, 0.0]
                elif model_type == 'bruggeman':
                    num_p = 0
                    default_bounds = []
                    default_initial = []
                else:
                    raise ValueError(f"Modelo de dispersión no soportado: {model_type}")
                    
                if bounds is None:
                    bounds = default_bounds
                if initial is None:
                    initial = default_initial
                    
                disp_layers.append({
                    'layer_idx': idx,
                    'model_type': model_type,
                    'param_bounds': bounds,
                    'param_initial': initial,
                    'num_params': num_p,
                    'start_idx': total_disp_params
                })
                total_disp_params += num_p
                
    if is_parametric:
        print(f"Modelos paramétricos activos: {[d['model_type'] for d in disp_layers]} | Total parámetros de dispersión: {total_disp_params}")
    
    # 2. ANALIZAR ESPESORES OPTIMIZABLES Y FIJOS
    is_optimizable = []
    opt_bounds = []
    fixed_values = {}
    
    for idx, b in enumerate(d_bounds):
        if isinstance(b, (tuple, list)):
            d_min, d_max = float(b[0]), float(b[1])
            if d_min == d_max:
                is_optimizable.append(False)
                fixed_values[idx] = d_min
            else:
                is_optimizable.append(True)
                opt_bounds.append((d_min, d_max))
        else:
            is_optimizable.append(False)
            fixed_values[idx] = float(b)
            
    num_opt_layers = sum(is_optimizable)
    print(f"Capas finitas a optimizar espesor: {num_opt_layers} | Capas de espesor fijo: {num_finite_layers - num_opt_layers}")
    
    # 3. INICIALIZACIÓN MULTI-START DE ESPESORES EN LOGITS
    d_initial = torch.zeros((num_starts, num_opt_layers), dtype=torch.float64, device=device)
    o_idx = 0
    for idx, opt in enumerate(is_optimizable):
        if opt:
            d_min, d_max = opt_bounds[o_idx]
            d_init_phys = torch.empty(num_starts, device=device).uniform_(d_min, d_max)
            p = (d_init_phys - d_min) / (d_max - d_min)
            p = torch.clamp(p, 1e-7, 1.0 - 1e-7)
            d_initial[:, o_idx] = torch.log(p / (1.0 - p))
            o_idx += 1
            
    d_opt = d_initial.clone().detach().requires_grad_(True)
    
    # 4. INICIALIZACIÓN MULTI-START DE PARÁMETROS DE DISPERSIÓN EN LOGITS
    if is_parametric:
        p_initial = torch.zeros((num_starts, total_disp_params), dtype=torch.float64, device=device)
        for d_lay in disp_layers:
            start_idx = d_lay['start_idx']
            bounds = d_lay['param_bounds']
            initial = d_lay['param_initial']
            
            for k in range(d_lay['num_params']):
                p_min, p_max = float(bounds[k][0]), float(bounds[k][1])
                # Semilla inicial en un entorno alrededor de initial
                init_val = float(initial[k])
                # Unificar aleatorización dentro de límites
                p_init_phys = torch.empty(num_starts, device=device).uniform_(p_min, p_max)
                p = (p_init_phys - p_min) / (p_max - p_min)
                p = torch.clamp(p, 1e-7, 1.0 - 1e-7)
                p_initial[:, start_idx + k] = torch.log(p / (1.0 - p))
                
        p_opt = p_initial.clone().detach().requires_grad_(True)
        optimizer = optim.Adam([d_opt, p_opt], lr=lr)
    else:
        optimizer = optim.Adam([d_opt], lr=lr)
        
    print(f"Iniciando optimización elipsométrica con Autograd ({num_starts} semillas en paralelo)...")
    
    inf_col = torch.full((num_starts, 1), float('inf'), dtype=torch.float64, device=device)
    
    def reconstruct_d_fisico(d_opt_tensor):
        d_fisico_cols = []
        o_idx = 0
        for idx, opt in enumerate(is_optimizable):
            if opt:
                d_min, d_max = opt_bounds[o_idx]
                col = d_min + (d_max - d_min) * torch.sigmoid(d_opt_tensor[:, o_idx])
                d_fisico_cols.append(col)
                o_idx += 1
            else:
                val = fixed_values[idx]
                col = torch.full((d_opt_tensor.shape[0],), val, dtype=torch.float64, device=device)
                d_fisico_cols.append(col)
        return torch.stack(d_fisico_cols, dim=1)
        
    def reconstruct_n_list_batched(p_opt_tensor):
        # Construir una lista de tensores por capa (sin operaciones in-place)
        layer_tensors = [None] * num_layers
        
        # 1. Copiar capas estáticas (no paramétricas)
        for idx in range(num_layers):
            is_this_parametric = False
            for d_lay in disp_layers:
                if d_lay['layer_idx'] == idx:
                    is_this_parametric = True
                    break
            if not is_this_parametric:
                layer_tensors[idx] = n_list[idx].unsqueeze(0).expand(num_starts, -1)
                
        # 2. Calcular capas con modelos de dispersión (Cauchy, etc.)
        for d_lay in disp_layers:
            if d_lay['model_type'] in ['cauchy', 'cauchy_absorbent']:
                idx = d_lay['layer_idx']
                start_idx = d_lay['start_idx']
                bounds = d_lay['param_bounds']
                model_type = d_lay['model_type']
                
                p_phys_list = []
                for k in range(d_lay['num_params']):
                    p_min, p_max = float(bounds[k][0]), float(bounds[k][1])
                    p_logit = p_opt_tensor[:, start_idx + k]
                    p_phys = p_min + (p_max - p_min) * torch.sigmoid(p_logit)
                    p_phys_list.append(p_phys)
                    
                lams_2d = lams.unsqueeze(0)  # [1, num_wl]
                if model_type == 'cauchy':
                    A = p_phys_list[0].unsqueeze(1)
                    B = p_phys_list[1].unsqueeze(1)
                    C = p_phys_list[2].unsqueeze(1)
                    n_calc = A + B * 1e4 / (lams_2d ** 2) + C * 1e9 / (lams_2d ** 4)
                    n_complex = n_calc.to(torch.complex128)
                elif model_type == 'cauchy_absorbent':
                    A = p_phys_list[0].unsqueeze(1)
                    B = p_phys_list[1].unsqueeze(1)
                    C = p_phys_list[2].unsqueeze(1)
                    D = p_phys_list[3].unsqueeze(1)
                    E = p_phys_list[4].unsqueeze(1)
                    F = p_phys_list[5].unsqueeze(1)
                    n_calc = A + B * 1e4 / (lams_2d ** 2) + C * 1e9 / (lams_2d ** 4)
                    k_calc = D + E * 1e4 / (lams_2d ** 2) + F * 1e9 / (lams_2d ** 4)
                    k_calc = torch.clamp(k_calc, min=0.0)
                    n_complex = torch.complex(n_calc, k_calc)
                    
                layer_tensors[idx] = n_complex

        # 3. Capas dependientes (Bruggeman EMA) - usan el resultado de Cauchy
        for d_lay in disp_layers:
            if d_lay['model_type'] == 'bruggeman':
                idx = d_lay['layer_idx']
                # Tomamos el tensor de la capa adyacente (idx + 1), ya calculada
                n_base = layer_tensors[idx + 1]
                f_air = 0.5
                # Tensor de aire (n=1.0) con la misma forma que n_base
                n_air = torch.ones_like(n_base)
                # Bruggeman directamente sobre tensores (mantiene Autograd)
                layer_tensors[idx] = n_eff_torch(n_base, n_air, f_air)

        # Apilar en tensor 3D sin operaciones in-place
        return torch.stack(layer_tensors, dim=1)

    # 5. BUCLE DE OPTIMIZACIÓN
    for epoch in range(num_epochs):
        optimizer.zero_grad()
        
        d_fisico = reconstruct_d_fisico(d_opt)
        d_full = torch.cat([inf_col, d_fisico, inf_col], dim=1)
        d_full_3d = d_full.unsqueeze(-1).expand(-1, -1, num_wl)
        
        if is_parametric:
            n_list_3d = reconstruct_n_list_batched(p_opt)
            res_s = coh_tmm_torch_batched(pol='s', n_list=n_list_3d, d_list=d_full_3d, th_0=th_0, lam_vac=lams)
            res_p = coh_tmm_torch_batched(pol='p', n_list=n_list_3d, d_list=d_full_3d, th_0=th_0, lam_vac=lams)
        else:
            res_s = tmm.coh_tmm_torch(pol='s', n_list=n_list, d_list=d_full_3d, th_0=th_0, lam_vac=lams)
            res_p = tmm.coh_tmm_torch(pol='p', n_list=n_list, d_list=d_full_3d, th_0=th_0, lam_vac=lams)
            
        r_s = res_s['r']
        r_p = res_p['r']
        
        rho = torch.conj(r_p / r_s)
        psi_teo = torch.atan(torch.abs(rho))
        delta_teo = torch.angle(rho)
        
        Is_teo = torch.sin(2 * psi_teo) * torch.sin(delta_teo)
        Ic_teo = torch.sin(2 * psi_teo) * torch.cos(delta_teo)
        
        error_Is = (Is_teo - Is_exp) ** 2
        error_Ic = (Ic_teo - Ic_exp) ** 2
        
        loss_per_start = (error_Is + error_Ic).mean(dim=-1)
        loss = loss_per_start.sum()
        
        loss.backward()
        optimizer.step()
        
        # Prevenir saturación de sigmoid: clampear logits para que
        # sigmoid se mantenga en [0.0067, 0.9933] y los gradientes no se anulen
        with torch.no_grad():
            d_opt.data.clamp_(-5, 5)
            if is_parametric:
                p_opt.data.clamp_(-5, 5)
        
        if epoch % 50 == 0 or epoch == num_epochs - 1:
            mejor_loss = loss_per_start.min().item()
            print(f"Epoch {epoch}/{num_epochs} | Mejor Loss (MSE): {mejor_loss:.6f}")
            
    # 6. EXTRACCIÓN DEL MÍNIMO GLOBAL
    with torch.no_grad():
        d_final_fisico = reconstruct_d_fisico(d_opt)
        d_full = torch.cat([inf_col, d_final_fisico, inf_col], dim=1)
        d_full_3d = d_full.unsqueeze(-1).expand(-1, -1, num_wl)
        
        if is_parametric:
            n_list_3d_final = reconstruct_n_list_batched(p_opt)
            res_s_final = coh_tmm_torch_batched(pol='s', n_list=n_list_3d_final, d_list=d_full_3d, th_0=th_0, lam_vac=lams)
            res_p_final = coh_tmm_torch_batched(pol='p', n_list=n_list_3d_final, d_list=d_full_3d, th_0=th_0, lam_vac=lams)
        else:
            res_s_final = tmm.coh_tmm_torch(pol='s', n_list=n_list, d_list=d_full_3d, th_0=th_0, lam_vac=lams)
            res_p_final = tmm.coh_tmm_torch(pol='p', n_list=n_list, d_list=d_full_3d, th_0=th_0, lam_vac=lams)
            
        r_s_final = res_s_final['r']
        r_p_final = res_p_final['r']
        
        rho_final = torch.conj(r_p_final / r_s_final)
        psi_teo_final = torch.atan(torch.abs(rho_final))
        delta_teo_final = torch.angle(rho_final)
        
        Is_teo_final = torch.sin(2 * psi_teo_final) * torch.sin(delta_teo_final)
        Ic_teo_final = torch.sin(2 * psi_teo_final) * torch.cos(delta_teo_final)
        
        error_Is_final = (Is_teo_final - Is_exp) ** 2
        error_Ic_final = (Ic_teo_final - Ic_exp) ** 2
        loss_final = (error_Is_final + error_Ic_final).mean(dim=-1)
        
    best_idx = loss_final.argmin()
    
    best_thicknesses = d_final_fisico[best_idx].cpu().detach().numpy()
    best_Is_curve = Is_teo_final[best_idx].cpu().detach().numpy()
    best_Ic_curve = Ic_teo_final[best_idx].cpu().detach().numpy()
    
    best_params = {}
    if is_parametric:
        with torch.no_grad():
            for d_lay in disp_layers:
                idx = d_lay['layer_idx']
                start_idx = d_lay['start_idx']
                bounds = d_lay['param_bounds']
                model_type = d_lay['model_type']
                
                if layer_names is not None and idx < len(layer_names):
                    layer_name = layer_names[idx]
                else:
                    layer_name = f"layer_{idx}"
                    
                p_best_phys = []
                for k in range(d_lay['num_params']):
                    p_min, p_max = float(bounds[k][0]), float(bounds[k][1])
                    p_logit = p_opt[best_idx, start_idx + k]
                    p_phys = p_min + (p_max - p_min) * torch.sigmoid(p_logit)
                    p_best_phys.append(p_phys.item())
                    
                if model_type == 'cauchy':
                    best_params[layer_name] = {
                        'model': 'cauchy',
                        'A': p_best_phys[0],
                        'B': p_best_phys[1],
                        'C': p_best_phys[2]
                    }
                elif model_type == 'cauchy_absorbent':
                    best_params[layer_name] = {
                        'model': 'cauchy_absorbent',
                        'A': p_best_phys[0],
                        'B': p_best_phys[1],
                        'C': p_best_phys[2],
                        'D': p_best_phys[3],
                        'E': p_best_phys[4],
                        'F': p_best_phys[5]
                    }
                elif model_type == 'bruggeman':
                    best_params[layer_name] = {
                        'model': 'bruggeman_EMA',
                        'linked_to': layer_names[idx+1] if layer_names else f'layer_{idx+1}',
                        'f_air': 0.5
                    }
                    
    print(f"\n¡Optimización elipsométrica completada!")
    print(f"Espesores óptimos encontrados: {best_thicknesses} nm")
    print(f"Error mínimo (MSE): {loss_final[best_idx].item():.6f}")
    if is_parametric:
        print("Parámetros de dispersión óptimos:")
        for name, params in best_params.items():
            print(f"  Material: {name} ({params['model']})")
            for k, val in params.items():
                if k != 'model':
                    if isinstance(val, (int, float)):
                        print(f"    {k} = {val:.6f}")
                    else:
                        print(f"    {k} = {val}")
                    
    if layer_models is not None:
        return best_thicknesses, best_Is_curve, best_Ic_curve, best_params
    else:
        return best_thicknesses, best_Is_curve, best_Ic_curve

def load_stack(filename, materials):
    '''Loads stack from Xlsx used by IR-S in Matlab. 'Materials' is a dict 
    where new materials will be added. Only materials not already present will 
    be added. YOU SHOULD CHECK that different nk's have different material names
    '''
    
    wb = load_workbook(filename = filename)
    ws = wb['MJSC Definition']
    
    stack = []
    start_row = 19
    end_row = 35
    
    for row in range(start_row, end_row+1):
        
        mat = ws.cell(row=row, column=3).value #C
        nk = ws.cell(row=row, column=5).value  #E
        thick = ws.cell(row=row, column=7).value #G
        
        if mat not in materials:
            # print('working on '+mat)
            materials[mat] = (load_interp('./nkdata/optical/' + nk))
        
        stack.append([thick, mat, 'c'])

    return  stack, materials

def plot_js(titulo, e_porosa, e_densa, js,espesores_comparación = None,label_comparación = "experimental",label_simulado = "optimo simulado"):
    print(f'caso: {titulo}')
    X, Y = np.meshgrid(e_porosa, e_densa)
    maxj = js.max()
    maxind = np.unravel_index(js.argmax(),js.shape)
    x_opt = X[maxind]
    y_opt = Y[maxind]
    print(f'e_superior: {x_opt}, e_inferior:{y_opt}')
    print(f'Max Jsc = {maxj:.4f}\n')
    Z = 100*js/maxj
    

    fig,ax=plt.subplots(1,1)
    colors =['blue','navy','indigo','purple','red','orangered','orange','w']
    cp = ax.contourf(X, Y, Z,levels=[70,80,85,90,95,98,99, 100], colors=colors)
    cbar =fig.colorbar(cp) # Add a colorbar to a plot
    # 8/4 le agregue este label
    cbar.set_label('Jsc_Normalizada [%]', rotation=270, labelpad=15)
    # 8/4 le agregue este scatter
    ax.scatter(x_opt, y_opt, color='white', edgecolors='black', s=100, marker='*', label=f'{label_simulado}')
    #ahora el valor experimental de Simon
    if espesores_comparación != None: 
        espesores_Simon = espesores_comparación
        ax.scatter(*espesores_Simon, color='green', edgecolors='black', s=100, marker='*', label=f'{label_comparación}')
    
    ax.set_title(titulo)
    ax.set_xlabel('Espesor superior [nm]')
    ax.set_ylabel('Espesor inferior [nm]')
    ax.legend()
    plt.show()
    
